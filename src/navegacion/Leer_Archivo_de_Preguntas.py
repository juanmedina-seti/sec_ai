"""
A web page with streamlit that upload an excel file, ask the column to read and the column to write in.
then allow the user got row by row reading from the column to read from. call a function that return 3 strings (answer1, answer2, generated)
show the tree string in and allows the user to edit them and select one of then to write in the write column.
Update the excel and continue moving to the next row
please do not include the langchain and chroma functions in this file.
"""
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from src.utils.answer_security_questions import previous_answers, new_answer
import os
load_dotenv()  # Load the .env file
temp_dir = os.environ.get("TEMP_DIR")

# Define the stages of the application
START = "START"
NAVEGATING = "NAVEGATING"
EDITING = "EDITING"


# Define the default stage
st.session_state["stage"] = START if "stage" not in st.session_state else st.session_state["stage"]
st.session_state["row_num"] = 0 if "row_num" not in st.session_state else st.session_state["row_num"]
# Define the function to move to the editing stage
def go_navegating():
    st.session_state["stage"] = NAVEGATING
    st.session_state["answer_1"] = None
    st.session_state["answer_2"] = None
    st.session_state["generated"] = None

# Define the function to move to the editing stage
def go_editing():
    st.session_state["stage"] = EDITING

def go_start():
    st.session_state["stage"] = START
    st.session_state["FILE"] = None
    st.session_state["selected_worksheet"] = None
    st.session_state["column_to_read_index"] = None
    st.session_state["column_to_write_index"] = None
    st.session_state["row_num"] = 0
    st.session_state["answer_1"] = None
    st.session_state["answer_2"] = None
    st.session_state["generated"] = None

# Define the function to load the excel file
def load_excel_file(file_name):
    excel_file = os.path.join(temp_dir, file_name)
    workbook = load_workbook(excel_file)
    return workbook

# Define the function to display the excel file
def display_excel_file(file_name,workbook, rows=10) -> tuple:
    selected_worksheet = st.selectbox("seleccione la hoja", [""] + workbook.sheetnames)
    column_to_read_index = 0
    column_to_write_index = 0
    excel_file = os.path.join(temp_dir, file_name)
    if selected_worksheet is not None and len(selected_worksheet) > 0:
        df = pd.read_excel(excel_file, sheet_name=selected_worksheet, nrows=rows+1)
        st.dataframe(df, height=300)
        columns_list = list([""] + list(df.columns))
        column_to_read, column_to_write = select_columns(columns_list)
        if column_to_read is not None and column_to_write is not None:
            column_to_read_index = columns_list.index(column_to_read)
            column_to_write_index = columns_list.index(column_to_write) 
    return  selected_worksheet, column_to_read_index,column_to_write_index


# Define the function to select the columns
def select_columns(column_names):
    column_to_read = st.selectbox("Selecciona la columna a leer", column_names)
    column_to_write = st.selectbox("Selecciona la columna a escribir", column_names)
    return column_to_read, column_to_write

# Define the function to edit the answers
def edit_answers(excel_file, selected_worksheet, column_to_read_index, column_to_write_index):
    workbook = load_workbook(excel_file)
    ws = workbook[selected_worksheet]
    row_num = st.session_state["row_num"] if "row_num" in st.session_state else 1
    
    st.header(f"Editando")
    st.write(f"Archivo: {excel_file}")
    st.write(f"Hoja: {selected_worksheet}")
    st.write(f"Preguntas columna: {column_to_read_index}, comentarios: {column_to_write_index}")
    st.write(f"Fila: {row_num}")

    row_content = [cell.value for cell in ws[row_num+1]]
    cell_question = ws.cell(row=row_num+1, column=column_to_read_index)
    cell_comments = ws.cell(row=row_num+1, column=column_to_write_index)

    #df = pd.DataFrame(data=row_content)
    #st.table(df)
    st.text_area("Pregunta o condición", value=str(cell_question.value), disabled=True)
    st.text_area("Respuesta", value=str(cell_comments.value),disabled=True )

    
    if st.button("Buscar") or st.session_state["stage"] == EDITING :
        st.session_state["stage"] == EDITING
        st.write("Respuestas anteriores")        
        
        if st.session_state["answer_1"] is None or st.session_state["answer_2"] is None:
            results = previous_answers(str(cell_question.value))
            answer_1 = results[0]
            answer_2 = results[1]
            st.session_state["answer_1"] = answer_1
            st.session_state["answer_2"] = answer_2

            generated = new_answer(str(cell_question.value), results=results)
            st.session_state["generated"] = generated
        else:       
            answer_1 = st.session_state["answer_1"]
            answer_2 = st.session_state["answer_2"]
            generated = st.session_state["generated"]


        col1, col2 = st.columns(2)
        edited_answer1 = col1.text_area("Respuesta 1", value=answer_1)
        edited_answer2 = col2.text_area("Respuesta 2", value=answer_2)
        st.write("Respuesta generada")
        edited_generated = st.text_area("Nueva Respuesta", value=generated)

        selected_answer = st.selectbox("Selecciona la respuesta a guardar", [
            "Nueva Respuesta", "Respuesta 1", "Respuesta 2", ])

        if selected_answer == "Respuesta 1":
            saved_answer = edited_answer1
        elif selected_answer == "Respuesta 2":
            saved_answer = edited_answer2
        else:
            saved_answer = edited_generated
        if (st.button("Guardar")):
            cell_comments.value = saved_answer
            workbook.save(excel_file)
            go_navegating()

    if  row_num > 1:
        if (st.button("<< Anterior")):
                st.session_state["row_num"] = row_num - 1
                go_navegating()

    if (st.button("Siguiente >>")):
            st.session_state["row_num"] = row_num + 1
            go_navegating()


# Define the function to edit the answers   

# Main application logic
st.title("Responder encuestas")

if st.session_state["stage"] == START:
    # Upload the Excel file
    filenames = os.listdir(temp_dir)
    excel_files = [""] + [x for x in filenames if x.endswith(".xlsx")]
    selected_filename = st.selectbox('Seleccione el archivo', excel_files)

    with st.popover("Subir archivo"):
        uploaded_file = st.file_uploader("Sube el archivo Excel", type=["xlsx"])
        if uploaded_file is not None:
            excel_file = os.path.join(temp_dir, uploaded_file.name)
            with open(excel_file, "wb") as f:
                f.write(uploaded_file.getbuffer())
            st.session_state["uploaded"] = uploaded_file.name

    if selected_filename is not None and len(selected_filename) > 0:
        workbook = load_excel_file(selected_filename)
        st.session_state["FILE"] = selected_filename
        selected_worksheet,column_to_read_index,column_to_write_index = display_excel_file(selected_filename,workbook)
        st.session_state["selected_worksheet"] = selected_worksheet
        st.session_state["column_to_read_index"] = column_to_read_index
        st.session_state["column_to_write_index"] = column_to_write_index
        workbook.close()
        st.button("Continuar", on_click=go_navegating)

elif st.session_state["stage"] != START:
    file_name = st.session_state["FILE"]
    excel_file = os.path.join(temp_dir, file_name)
    selected_worksheet = st.session_state["selected_worksheet"]
    column_to_read_index = st.session_state["column_to_read_index"]
    column_to_write_index = st.session_state["column_to_write_index"]
    edit_answers(excel_file, selected_worksheet, column_to_read_index, column_to_write_index)



